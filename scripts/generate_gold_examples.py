import os
import re
import csv
import json
import docx
import openpyxl

MANIFEST_PATH = 'data/raw/manifest.csv'
SCRIPTS_DIR = 'data/raw/scripts'
STORYBOARDS_DIR = 'data/raw/storyboards'
OUTPUT_DIR = 'data/processed/json'

def get_module_id(script_file, storyboard_file):
    # Try to match Umed style first: Umed_2.1.1_Script.docx -> 2.1.1
    for f in (script_file, storyboard_file):
        if f:
            match = re.search(r'Umed_(\d+(?:\.\d+)+)_', f)
            if match:
                return match.group(1)
    
    # Try to match other styles: ADGM UBO Course_Proposed Modules_1.docx -> adgm_1
    for f in (script_file, storyboard_file):
        if f:
            if 'ADGM' in f.upper():
                match = re.search(r'(\d+)', f)
                if match:
                    return f"adgm_{match.group(1)}"
            match = re.search(r'_(\d+)\.(docx|xlsx)', f)
            if match:
                return match.group(1)
            # fallback: find first digit sequence
            match = re.search(r'(\d+)', f)
            if match:
                return match.group(1)
    return 'unknown'

def compute_source_text(beat):
    d = beat.get("dialogue", "").strip()
    n = beat.get("narration", "").strip()
    o = beat.get("ost", "").strip()

    if d:
        return beat["dialogue"]
    elif n:
        return beat["narration"]
    elif o:
        return beat["ost"]
    else:
        return ""


def normalize_cell(text):
    if text is None:
        return ""
    return re.sub(r'[^a-z]', '', str(text).lower())

def match_linear_header(headers):
    cleaned = [normalize_cell(h) for h in headers]
    if len(cleaned) >= 3:
        return (cleaned[0] == 'graphicsanimationtext' and 
                cleaned[1] == 'onscreen' and 
                cleaned[2] == 'narratordialog')
    return False

def match_interactive_header(headers):
    cleaned = [normalize_cell(h) for h in headers]
    if len(cleaned) >= 5:
        return (cleaned[0] == 'visual' and 
                cleaned[1] == 'interaction' and 
                cleaned[2] == 'ost' and 
                cleaned[3] == 'dialogue' and 
                cleaned[4] == 'narration')
    return False

def extract_beats_docx(filepath, dialect, module_id):
    doc = docx.Document(filepath)
    matched_table = None
    is_interactive_matched = False
    is_linear_matched = False

    for table in doc.tables:
        if not table.rows:
            continue
        headers = [c.text for c in table.rows[0].cells]
        
        if dialect == 'linear' and match_linear_header(headers):
            matched_table = table
            is_linear_matched = True
            break
        elif dialect == 'interactive' and match_interactive_header(headers):
            matched_table = table
            is_interactive_matched = True
            break
        # Fallback check both patterns just in case
        elif match_linear_header(headers):
            matched_table = table
            is_linear_matched = True
            break
        elif match_interactive_header(headers):
            matched_table = table
            is_interactive_matched = True
            break

    if matched_table is None:
        return None, "No table matching expected column headers found"

    beats = []
    # If the matched table headers didn't match the dialect specified in manifest, we warn but extract
    actual_dialect = 'interactive' if is_interactive_matched else 'linear'
    
    for i, row in enumerate(matched_table.rows[1:], start=1):
        cells = [c.text.strip() for c in row.cells]
        beat_id = f"{module_id}_beat_{i:02d}"
        
        if actual_dialect == 'linear':
            beat = {
                "beat_id": beat_id,
                "visual": cells[0] if len(cells) > 0 else "",
                "ost": cells[1] if len(cells) > 1 else "",
                "dialogue": cells[2] if len(cells) > 2 else ""
            }
        else:
            beat = {
                "beat_id": beat_id,
                "visual": cells[0] if len(cells) > 0 else "",
                "interaction": cells[1] if len(cells) > 1 else "",
                "ost": cells[2] if len(cells) > 2 else "",
                "dialogue": cells[3] if len(cells) > 3 else "",
                "narration": cells[4] if len(cells) > 4 else ""
            }
        beat["source_text"] = compute_source_text(beat)
        beats.append(beat)
        
    return beats, None

def extract_beats_xlsx(filepath, dialect, module_id):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    
    header_row_idx = None
    actual_dialect = None
    
    # Search for header row
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [str(c) if c is not None else "" for c in row]
        if match_interactive_header(headers):
            header_row_idx = r_idx
            actual_dialect = 'interactive'
            break
        elif match_linear_header(headers):
            header_row_idx = r_idx
            actual_dialect = 'linear'
            break
            
    if header_row_idx is None:
        return None, "No header row matching expected column patterns found"
        
    beats = []
    beat_num = 1
    
    # Read rows below header
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_values = [str(c).strip() if c is not None else "" for c in row]
        
        # Check if row is blank (using the columns of interest)
        limit = 5 if actual_dialect == 'interactive' else 3
        # Pad row_values if it's shorter than limit
        while len(row_values) < limit:
            row_values.append("")
            
        if all(val == "" for val in row_values[:limit]):
            break
            
        beat_id = f"{module_id}_beat_{beat_num:02d}"
        if actual_dialect == 'linear':
            beat = {
                "beat_id": beat_id,
                "visual": row_values[0],
                "ost": row_values[1],
                "dialogue": row_values[2]
            }
        else:
            beat = {
                "beat_id": beat_id,
                "visual": row_values[0],
                "interaction": row_values[1],
                "ost": row_values[2],
                "dialogue": row_values[3],
                "narration": row_values[4]
            }
        beat["source_text"] = compute_source_text(beat)
        beats.append(beat)
        beat_num += 1
        
    return beats, None

def run():
    # Save the original modification timestamp of manifest.csv
    orig_mtime = os.path.getmtime(MANIFEST_PATH)
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    attempted = 0
    successful = 0
    total_beats = 0
    failed_modules = []
    
    all_beats_flattened = []
    
    with open(MANIFEST_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            flag = row.get('flag', '').strip()
            if flag:
                # Skip incomplete rows
                continue
                
            attempted += 1
            script_file = row.get('script_file', '').strip()
            storyboard_file = row.get('storyboard_file', '').strip()
            dialect = row.get('dialect', '').strip().lower()
            client = row.get('client', '').strip()
            subject = row.get('subject', '').strip()
            audience = row.get('audience', '').strip()
            language = row.get('language', '').strip()
            
            module_id = get_module_id(script_file, storyboard_file)
            
            storyboard_path = os.path.join(STORYBOARDS_DIR, storyboard_file)
            if not os.path.exists(storyboard_path):
                failed_modules.append((module_id, f"Storyboard file not found: {storyboard_file}"))
                continue
                
            beats = None
            err = None
            if storyboard_file.endswith('.docx'):
                beats, err = extract_beats_docx(storyboard_path, dialect, module_id)
            elif storyboard_file.endswith('.xlsx'):
                beats, err = extract_beats_xlsx(storyboard_path, dialect, module_id)
            else:
                err = f"Unsupported storyboard file format: {storyboard_file}"
                
            if err:
                failed_modules.append((module_id, err))
                continue
                
            if not beats:
                failed_modules.append((module_id, "Zero beats extracted"))
                continue
                
            # Write per-module JSON file
            # replace any dots in module number with underscores in the filename
            file_module_id = module_id.replace('.', '_')
            output_json_path = os.path.join(OUTPUT_DIR, f"{file_module_id}.json")
            
            module_data = {
                "module_id": module_id,
                "client": client,
                "subject": subject,
                "audience": audience,
                "language": language,
                "dialect": dialect,
                "beats": beats
            }
            
            with open(output_json_path, 'w', encoding='utf-8') as out_f:
                json.dump(module_data, out_f, indent=2, ensure_ascii=False)
                
            successful += 1
            total_beats += len(beats)
            print(f"Successfully processed module {module_id} -> {output_json_path} ({len(beats)} beats)")
            
            # Prepare flattened beats for gold_examples.jsonl
            for b in beats:
                flat_beat = {
                    "beat_id": b["beat_id"],
                    "client": client,
                    "subject": subject,
                    "audience": audience,
                    "language": language,
                    "dialect": dialect,
                }
                # add the rest of keys from beat
                for k, v in b.items():
                    if k != "beat_id":
                        flat_beat[k] = v
                all_beats_flattened.append(flat_beat)

    # Write gold_examples.jsonl
    gold_jsonl_path = os.path.join(OUTPUT_DIR, "gold_examples.jsonl")
    with open(gold_jsonl_path, 'w', encoding='utf-8') as out_f:
        for fb in all_beats_flattened:
            out_f.write(json.dumps(fb, ensure_ascii=False) + '\n')
            
    empty_source_beats = [b for b in all_beats_flattened if not b.get("source_text")]

    print("\n================== SUMMARY ==================")
    print(f"Total modules attempted: {attempted}")
    print(f"Total modules successfully converted: {successful}")
    print(f"Total beats extracted across all modules: {total_beats}")
    print(f"Total beats with empty source_text: {len(empty_source_beats)}")

    if empty_source_beats:
        print("\nBeats with empty source_text (need manual review):")
        for eb in empty_source_beats:
            print(f"  - Beat ID: {eb['beat_id']} | Visual: {repr(eb.get('visual'))} | Interaction: {repr(eb.get('interaction'))}")
    
    if failed_modules:
        print("\nFailed modules:")
        for mod, reason in failed_modules:
            print(f"  - Module {mod}: {reason}")
    else:
        print("\nNo modules failed.")
        
    # Verify manifest.csv is unmodified
    new_mtime = os.path.getmtime(MANIFEST_PATH)
    if orig_mtime == new_mtime:
        print(f"\nConfirmation: {MANIFEST_PATH} last-modified timestamp is unchanged.")
    else:
        print(f"\nWARNING: {MANIFEST_PATH} last-modified timestamp changed!")

if __name__ == '__main__':
    run()
